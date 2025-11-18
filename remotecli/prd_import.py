from pathlib import Path
from typing import List, Dict
from openpyxl import load_workbook

def _norm(s: str) -> str:
    return s.strip().lower().replace(" ", "_")

def _get(row: dict, keys: List[str]) -> str:
    for k in keys:
        v = row.get(k)
        if v is not None:
            return str(v)
    return ""

def _split_rules(v: str) -> List[str]:
    if not v:
        return []
    parts = [p.strip() for p in str(v).replace(";", ",").split(",")]
    return [p for p in parts if p]

def load_prd_xlsx(file_path: str) -> List[Dict]:
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    wb = load_workbook(filename=str(p), data_only=True)
    items: List[Dict] = []
    for ws in wb.worksheets:
        headers = []
        for cell in ws[1]:
            headers.append(_norm(str(cell.value or "")))
        idx = {h: i for i, h in enumerate(headers)}
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for h, i in idx.items():
                d[h] = r[i] if i is not None and i < len(r) else None
            rows.append(d)
        for row in rows:
            osv = _get(row, ["os", "platform"])
            cid = _get(row, ["control_id", "control", "controlid"])
            desc = _get(row, ["description", "desc"])
            rid = _get(row, ["rule_id", "rule", "ruleid"])
            rids = _split_rules(rid or _get(row, ["rule_ids", "rules"]))
            if not (osv and cid):
                continue
            items.append({"os": osv.lower(), "control_id": cid, "description": desc, "rule_ids": rids})
    return items